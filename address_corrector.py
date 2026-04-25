"""
Address Corrector
-----------------
Reads a CSV or Excel file containing address fields, applies standardisation
corrections, and writes an output file with both the original and corrected
columns side-by-side.

Usage:
    python address_corrector.py <input_file> <output_file>

Supported formats:  .csv  .xlsx  .xls
"""

import os
import re
import sys
from difflib import get_close_matches

import pandas as pd
import pycountry
import geonamescache as _gnc
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# COLUMN NAME DETECTION
# ──────────────────────────────────────────────────────────────────────────────

# ──────────────────────────────────────────────────────────────────────────────
# COLUMN DETECTION  — keyword-based, not a fixed list
# Works with any column name that contains relevant keywords or digits
# ──────────────────────────────────────────────────────────────────────────────

# Keywords that must appear somewhere in the column name (after normalisation)
# Listed in priority order for each field; first hit wins.
_FIELD_KEYWORDS = {
    # Address lines — distinguished by trailing digit (1 / 2 / 3)
    "address_line_1": [
        r"\baddr(?:ess)?\b.*\b1\b",
        r"\bstreet\b.*\b1\b",
        r"\bline\b.*\b1\b",
        r"\baddress1\b",
        r"\baddr1\b",
        r"\badd1\b",
        r"\bhouse\s*(?:no\.?|number|num)?\b",
        # Catch bare "address" / "street address" only when no digit variant exists (resolved later)
        r"\b(?:mailing|billing|delivery|residential|primary|house|full)?\s*addr(?:ess)?\b",
        r"\bstreet\s+addr(?:ess)?\b",
        r"\bstreet\b",
    ],
    "address_line_2": [
        r"\baddr(?:ess)?\b.*\b2\b",
        r"\bstreet\b.*\b2\b",
        r"\bline\b.*\b2\b",
        r"\baddress2\b",
        r"\baddr2\b",
        r"\badd2\b",
        r"\b(?:suite|apartment|unit|apt|floor|building)\b",
        r"\bsubpremise\b",
    ],
    "address_line_3": [
        r"\baddr(?:ess)?\b.*\b3\b",
        r"\bstreet\b.*\b3\b",
        r"\bline\b.*\b3\b",
        r"\baddress3\b",
        r"\baddr3\b",
        r"\badd3\b",
        r"\badditional\s*addr(?:ess)?\b",
        r"\bextra\s*addr(?:ess)?\b",
        r"\bsupplementary\s*addr(?:ess)?\b",
        r"\baddr(?:ess)?\s*(?:cont(?:inued)?|cont\.?)\b",
    ],
    "city": [
        r"\bcity\b",
        r"\btown\b",
        r"\bsuburb\b",
        r"\blocality\b",
        r"\bmunicipality\b",
        r"\btownship\b",
        r"\bvillage\b",
    ],
    "state": [
        r"\bstate\b",
        r"\bprovince\b",
        r"\bregion\b",
        r"\bcounty\b",
        r"\bterritory\b",
        r"\bprefecture\b",
    ],
    "country": [
        r"\bcountry\b",
        r"\bnation\b",
        r"\bnationality\b",
    ],
    "company_name": [
        r"\bname\b",
        r"\bcompany\b",
        r"\bvendor\s*name\b",
        r"\bclient\s*name\b",
        r"\bcustomer\s*name\b",
        r"\borganis?ation\b",
        r"\bfirm\b",
        r"\bbusiness\s*name\b",
        r"\bentity\b",
    ],
    "postal_code": [
        r"\bpostal\b",
        r"\bpostalcode\b",
        r"postal.?code",
        r"\bpostcode\b",
        r"\bzip\b",
        r"\bpin\s*code\b",
        r"\bpincode\b",
        r"\bpin\b",
        r"\bpost\s*code\b",
    ],
}


def _normalise_col_name(name):
    """Lower-case, collapse spaces/underscores/hyphens."""
    return re.sub(r"[\s_\-/]+", " ", name.lower().strip())


def _detect_columns(df_cols):
    """
    Map each address field to the best matching column name (or None).

    Strategy:
      1. For each column, test patterns in priority order.
      2. When multiple columns match the same field, pick the best one
         (highest priority pattern index = lower number wins).
      3. Each column can only be assigned to one field.
    """
    normalised = {c: _normalise_col_name(c) for c in df_cols}

    # candidates[field] = list of (priority, original_col_name)
    candidates = {f: [] for f in _FIELD_KEYWORDS}

    for orig_col, norm in normalised.items():
        for field, patterns in _FIELD_KEYWORDS.items():
            for priority, pat in enumerate(patterns):
                if re.search(pat, norm):
                    candidates[field].append((priority, orig_col))
                    break  # only record lowest-priority match per field per column

    # Resolve: assign each column to at most one field, prefer highest priority
    assigned_cols = set()
    col_map = {}

    for field in _FIELD_KEYWORDS:
        best = sorted(candidates[field], key=lambda x: x[0])
        matched = None
        for _, col in best:
            if col not in assigned_cols:
                matched = col
                assigned_cols.add(col)
                break
        col_map[field] = matched

    return col_map


# ──────────────────────────────────────────────────────────────────────────────
# LOOKUP TABLES
# ──────────────────────────────────────────────────────────────────────────────

# US states (full name → 2-letter abbreviation)
US_STATES = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "florida": "FL", "georgia": "GA", "hawaii": "HI", "idaho": "ID",
    "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN", "mississippi": "MS",
    "missouri": "MO", "montana": "MT", "nebraska": "NE", "nevada": "NV",
    "new hampshire": "NH", "new jersey": "NJ", "new mexico": "NM", "new york": "NY",
    "north carolina": "NC", "north dakota": "ND", "ohio": "OH", "oklahoma": "OK",
    "oregon": "OR", "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT",
    "vermont": "VT", "virginia": "VA", "washington": "WA", "west virginia": "WV",
    "wisconsin": "WI", "wyoming": "WY", "district of columbia": "DC",
    "puerto rico": "PR", "guam": "GU", "virgin islands": "VI",
    "american samoa": "AS", "northern mariana islands": "MP",
}

# Valid 2-letter US state codes (for uppercase normalisation)
US_STATE_CODES = set(US_STATES.values())

# Canadian provinces/territories (full name → 2-letter abbreviation)
CA_PROVINCES = {
    "alberta": "AB", "british columbia": "BC", "manitoba": "MB",
    "new brunswick": "NB", "newfoundland and labrador": "NL",
    "newfoundland": "NL", "labrador": "NL",
    "northwest territories": "NT", "nova scotia": "NS", "nunavut": "NU",
    "ontario": "ON", "prince edward island": "PE", "quebec": "QC",
    "québec": "QC", "saskatchewan": "SK", "yukon": "YT",
}

CA_PROVINCE_CODES = set(CA_PROVINCES.values())

# Australian states/territories
AU_STATES = {
    "new south wales": "NSW", "victoria": "VIC", "queensland": "QLD",
    "south australia": "SA", "western australia": "WA", "tasmania": "TAS",
    "northern territory": "NT", "australian capital territory": "ACT",
}

AU_STATE_CODES = set(AU_STATES.values())

# All known 2-letter region codes (do not title-case these)
ALL_REGION_CODES = US_STATE_CODES | CA_PROVINCE_CODES | AU_STATE_CODES | {
    # Indian states abbreviations kept as entered, just uppercase
    "AP", "AR", "AS", "BR", "CG", "GA", "GJ", "HR", "HP", "JH", "KA",
    "KL", "MP", "MH", "MN", "ML", "MZ", "NL", "OD", "PB", "RJ", "SK",
    "TN", "TS", "TR", "UP", "UK", "WB", "AN", "CH", "DN", "DD", "DL",
    "JK", "LA", "LD", "PY",
}

# Country aliases → standardised name
# Manual alias map — covers abbreviations, native-language names, and
# common alternate spellings that pycountry cannot resolve on its own.
# pycountry handles the remaining 200+ countries automatically.
COUNTRY_MAP = {
    # ── United States ──────────────────────────────────────────────────────
    "usa": "US", "u.s.a": "US", "u.s.a.": "US",
    "us": "US", "u.s.": "US", "u.s": "US",
    "united states of america": "US", "america": "US",
    "the united states": "US", "the us": "US",
    # ── United Kingdom ─────────────────────────────────────────────────────
    "uk": "GB", "u.k.": "GB", "u.k": "GB",
    "great britain": "GB", "gb": "GB", "g.b.": "GB",
    "england": "GB", "britain": "GB",
    "scotland": "GB", "wales": "GB",
    "northern ireland": "GB",
    # ── UAE ────────────────────────────────────────────────────────────────
    "uae": "AE", "u.a.e.": "AE",
    "u.a.e": "AE", "emirates": "AE",
    # ── Germany ────────────────────────────────────────────────────────────
    "deutschland": "DE", "allemagne": "DE", "almanya": "DE",
    # ── France ─────────────────────────────────────────────────────────────
    "frankreich": "FR", "frankrig": "FR",
    # ── Italy ──────────────────────────────────────────────────────────────
    "italia": "IT", "italie": "IT", "italien": "IT",
    # ── Spain ──────────────────────────────────────────────────────────────
    "espana": "ES", "españa": "ES", "espagne": "ES", "spanien": "ES",
    # ── Netherlands ────────────────────────────────────────────────────────
    "the netherlands": "NL", "holland": "NL",
    "nederland": "NL", "niederlande": "NL",
    # ── Belgium ────────────────────────────────────────────────────────────
    "belgique": "BE", "belgien": "BE", "belgië": "BE",
    # ── Switzerland ────────────────────────────────────────────────────────
    "schweiz": "CH", "suisse": "CH", "svizzera": "CH",
    "confederazione svizzera": "CH",
    # ── Sweden ─────────────────────────────────────────────────────────────
    "sverige": "SE", "schweden": "SE",
    # ── Norway ─────────────────────────────────────────────────────────────
    "norge": "NO", "norwegen": "NO",
    # ── Denmark ────────────────────────────────────────────────────────────
    "danmark": "DK", "dänemark": "DK",
    # ── Finland ────────────────────────────────────────────────────────────
    "suomi": "FI", "finnland": "FI",
    # ── Austria ────────────────────────────────────────────────────────────
    "österreich": "AT", "osterreich": "AT", "autriche": "AT",
    # ── Poland ─────────────────────────────────────────────────────────────
    "polska": "PL", "pologne": "PL", "polen": "PL",
    # ── Czech Republic ─────────────────────────────────────────────────────
    "czechia": "CZ", "ceska republika": "CZ",
    "česká republika": "CZ", "tschechien": "CZ",
    # ── Hungary ────────────────────────────────────────────────────────────
    "magyarország": "HU", "magyarorszag": "HU", "hongrie": "HU",
    # ── Greece ─────────────────────────────────────────────────────────────
    "hellas": "GR", "ellada": "GR", "griechenland": "GR",
    # ── Russia ─────────────────────────────────────────────────────────────
    "russian federation": "RU", "rossiya": "RU", "russland": "RU",
    # ── Turkey ─────────────────────────────────────────────────────────────
    "türkiye": "TR", "turkiye": "TR", "türkei": "TR",
    # ── Saudi Arabia ───────────────────────────────────────────────────────
    "ksa": "SA", "kingdom of saudi arabia": "SA",
    "al-mamlaka al-arabiyya as-saudiyya": "SA",
    # ── China ──────────────────────────────────────────────────────────────
    "prc": "CN", "people's republic of china": "CN",
    "peoples republic of china": "CN", "zhongguo": "CN",
    # ── South Korea ────────────────────────────────────────────────────────
    "korea": "KR", "republic of korea": "KR",
    "south korea": "KR", "hanguk": "KR",
    # ── North Korea ────────────────────────────────────────────────────────
    "dprk": "KP", "north korea": "KP",
    # ── Taiwan ─────────────────────────────────────────────────────────────
    "taiwan": "TW", "roc": "TW", "republic of china": "TW",
    # ── Hong Kong ──────────────────────────────────────────────────────────
    "hong kong sar": "HK", "hksar": "HK",
    # ── Brazil ─────────────────────────────────────────────────────────────
    "brasil": "BR", "brésil": "BR",
    # ── Mexico ─────────────────────────────────────────────────────────────
    "méxico": "MX", "mejico": "MX",
    # ── Australia ──────────────────────────────────────────────────────────
    "oz": "AU", "aussie": "AU",
    # ── New Zealand ────────────────────────────────────────────────────────
    "aotearoa": "NZ",
    # ── South Africa ───────────────────────────────────────────────────────
    "rsa": "ZA", "suid-afrika": "ZA",
    # ── India ──────────────────────────────────────────────────────────────
    "bharat": "IN", "hindustan": "IN",
    # ── Pakistan ───────────────────────────────────────────────────────────
    "pak": "PK",
    # ── Iran ───────────────────────────────────────────────────────────────
    "persia": "IR", "islamic republic of iran": "IR",
    # ── Iraq ───────────────────────────────────────────────────────────────
    "al-iraq": "IQ",
    # ── Egypt ──────────────────────────────────────────────────────────────
    "misr": "EG", "arab republic of egypt": "EG",
    # ── Morocco ────────────────────────────────────────────────────────────
    "maroc": "MA", "marruecos": "MA", "marokko": "MA",
    # ── Algeria ────────────────────────────────────────────────────────────
    "algérie": "DZ", "algerie": "DZ",
    # ── Tunisia ────────────────────────────────────────────────────────────
    "tunisie": "TN",
    # ── Ethiopia ───────────────────────────────────────────────────────────
    "abyssinia": "ET",
    # ── Ivory Coast ────────────────────────────────────────────────────────
    "ivory coast": "CI", "cote d'ivoire": "CI",
    "cote divoire": "CI",
    # ── Democratic Republic of Congo ───────────────────────────────────────
    "drc": "CD",
    "dr congo": "CD",
    "democratic republic of congo": "CD",
    "zaire": "CD",
    # ── Republic of Congo ──────────────────────────────────────────────────
    "congo": "CG",
    # ── Vietnam ────────────────────────────────────────────────────────────
    "viet nam": "VN",
    # ── Myanmar ────────────────────────────────────────────────────────────
    "burma": "MM",
    # ── Sri Lanka ──────────────────────────────────────────────────────────
    "ceylon": "LK",
    # ── Cambodia ───────────────────────────────────────────────────────────
    "kampuchea": "KH",
    # ── Bosnia ─────────────────────────────────────────────────────────────
    "bosnia": "BA", "bih": "BA",
    # ── Macedonia ──────────────────────────────────────────────────────────
    "north macedonia": "MK", "macedonia": "MK",
    # ── Kosovo ─────────────────────────────────────────────────────────────
    "kosovo": "XK",
    # ── Palestine ──────────────────────────────────────────────────────────
    "palestine": "PS",
    # ── Vatican ────────────────────────────────────────────────────────────
    "vatican": "VA", "vatican city": "VA",
    # ── Macau ──────────────────────────────────────────────────────────────
    "macau": "MO", "macao sar": "MO",
    # ── Singapore ──────────────────────────────────────────────────────────
    "singapore": "SG",
    # ── Japan ──────────────────────────────────────────────────────────────
    "japan": "JP",
    # ── Israel ─────────────────────────────────────────────────────────────
    "israel": "IL",
    # ── Romania ────────────────────────────────────────────────────────────
    "romania": "RO",
    # ── Ukraine ────────────────────────────────────────────────────────────
    "ukraine": "UA",
    # ── Bangladesh ─────────────────────────────────────────────────────────
    "bangladesh": "BD",
    # ── Indonesia ──────────────────────────────────────────────────────────
    "indonesia": "ID",
    # ── Malaysia ───────────────────────────────────────────────────────────
    "malaysia": "MY",
    # ── Philippines ────────────────────────────────────────────────────────
    "philippines": "PH",
    # ── Thailand ───────────────────────────────────────────────────────────
    "thailand": "TH",
    # ── Nigeria ────────────────────────────────────────────────────────────
    "nigeria": "NG",
    # ── Kenya ──────────────────────────────────────────────────────────────
    "kenya": "KE",
    # ── Ghana ──────────────────────────────────────────────────────────────
    "ghana": "GH",
    # ── Peru ───────────────────────────────────────────────────────────────
    "peru": "PE",
    # ── Venezuela ──────────────────────────────────────────────────────────
    "venezuela": "VE",
    # ── Ecuador ────────────────────────────────────────────────────────────
    "ecuador": "EC",
    # ── Bolivia ────────────────────────────────────────────────────────────
    "bolivia": "BO",
    # ── Paraguay ───────────────────────────────────────────────────────────
    "paraguay": "PY",
    # ── Uruguay ────────────────────────────────────────────────────────────
    "uruguay": "UY",
    # ── Cuba ───────────────────────────────────────────────────────────────
    "cuba": "CU",
    # ── Jamaica ────────────────────────────────────────────────────────────
    "jamaica": "JM",
    # ── Trinidad and Tobago ────────────────────────────────────────────────
    "trinidad and tobago": "TT",
    # ── Barbados ───────────────────────────────────────────────────────────
    "barbados": "BB",
    # ── Argentina ──────────────────────────────────────────────────────────
    "argentina": "AR",
    # ── Chile ──────────────────────────────────────────────────────────────
    "chile": "CL",
    # ── Colombia ───────────────────────────────────────────────────────────
    "colombia": "CO",
    # ── Portugal ───────────────────────────────────────────────────────────
    "portugal": "PT",
    # ── Ireland ────────────────────────────────────────────────────────────
    "ireland": "IE", "eire": "IE",
}

# ──────────────────────────────────────────────────────────────────────────────
# STREET TYPE ABBREVIATIONS  (USPS standard — always uppercase)
# ──────────────────────────────────────────────────────────────────────────────

STREET_TYPES = {
    "alley": "ALY", "aly": "ALY",
    "avenue": "AVE", "ave": "AVE", "av": "AVE",
    "boulevard": "BLVD", "blvd": "BLVD",
    "bypass": "BYP", "byp": "BYP",
    "causeway": "CSWY", "cswy": "CSWY",
    "circle": "CIR", "cir": "CIR",
    "close": "CL", "cl": "CL",
    "court": "CT", "ct": "CT",
    "cove": "CV", "cv": "CV",
    "crescent": "CRES", "cres": "CRES",
    "crossing": "XING", "xing": "XING",
    "drive": "DR", "dr": "DR",
    "estate": "EST", "est": "EST",
    "expressway": "EXPY", "expy": "EXPY",
    "extension": "EXT", "ext": "EXT",
    "freeway": "FWY", "fwy": "FWY",
    "grove": "GRV", "grv": "GRV",
    "heights": "HTS", "hts": "HTS",
    "highway": "HWY", "hwy": "HWY",
    "hill": "HL", "hl": "HL",
    "hollow": "HOLW", "holw": "HOLW",
    "junction": "JCT", "jct": "JCT",
    "lake": "LK", "lk": "LK",
    "lane": "LN", "ln": "LN",
    "loop": "LOOP",
    "manor": "MNR", "mnr": "MNR",
    "meadow": "MDW", "mdw": "MDW",
    "mount": "MT", "mt": "MT",
    "motorway": "MTWY", "mtwy": "MTWY",
    "park": "PARK",
    "parkway": "PKWY", "pkwy": "PKWY", "pky": "PKWY",
    "place": "PL", "pl": "PL",
    "plaza": "PLZ", "plz": "PLZ",
    "point": "PT", "pt": "PT",
    "ridge": "RDG", "rdg": "RDG",
    "road": "RD", "rd": "RD",
    "row": "ROW",
    "run": "RUN",
    "square": "SQ", "sq": "SQ",
    "street": "ST", "str": "ST",
    "terrace": "TER", "ter": "TER", "tce": "TER",
    "trail": "TRL", "trl": "TRL",
    "turnpike": "TPKE", "tpke": "TPKE",
    "valley": "VLY", "vly": "VLY",
    "vista": "VIS", "vis": "VIS",
    "walk": "WALK",
    "way": "WAY",
}

# Directional prefixes/suffixes
DIRECTIONALS = {
    "north": "N", "south": "S", "east": "E", "west": "W",
    "northeast": "NE", "northwest": "NW", "southeast": "SE", "southwest": "SW",
    "n.": "N", "s.": "S", "e.": "E", "w.": "W",
    "n": "N", "s": "S", "e": "E", "w": "W",
    "ne": "NE", "nw": "NW", "se": "SE", "sw": "SW",
    "n.e.": "NE", "n.w.": "NW", "s.e.": "SE", "s.w.": "SW",
}

# Secondary unit designators
UNIT_TYPES = {
    "apartment": "APT", "apt": "APT",
    "suite": "STE", "ste": "STE", "suit": "STE",
    "unit": "UNIT", "unt": "UNIT",
    "room": "RM", "rm": "RM",
    "floor": "FL", "fl": "FL", "flr": "FL",
    "building": "BLDG", "bldg": "BLDG", "bld": "BLDG",
    "department": "DEPT", "dept": "DEPT",
    "po box": "PO BOX", "p.o. box": "PO BOX", "p.o box": "PO BOX",
    "po. box": "PO BOX", "pobox": "PO BOX", "post office box": "PO BOX",
    "p o box": "PO BOX", "p.o. box": "PO BOX",
}

# ──────────────────────────────────────────────────────────────────────────────
# CORRECTION FUNCTIONS
# ──────────────────────────────────────────────────────────────────────────────

_NULL_PLACEHOLDERS = {
    "", "nan", "NaN", "None", "none", "null", "NULL", "Null",
    "<null>", "<Null>", "<NULL>", "<none>", "<None>", "<NONE>",
    "n/a", "N/A", "na", "NA", "<n/a>", "<N/A>",
    "nil", "NIL", "undefined", "UNDEFINED", "#N/A", "-",
}

def _clean(val):
    """Strip and collapse internal whitespace; treat null-like placeholders as empty."""
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s in _NULL_PLACEHOLDERS:
        return ""
    return re.sub(r"\s+", " ", s)


def _normalise_po_box(val):
    """Standardise PO Box variations to 'PO BOX <number>'."""
    pattern = re.compile(
        r"\b(p\.?\s*o\.?\s*box|post\s+office\s+box|pobox)\b\s*(\d*)",
        re.IGNORECASE,
    )
    return pattern.sub(lambda m: f"PO BOX {m.group(2)}".strip(), val)


def correct_address_line(val):
    val = _clean(val)
    if not val:
        return ""

    val = _normalise_po_box(val)

    words = val.split()
    result = []
    i = 0
    while i < len(words):
        w = words[i]
        wl = w.lower().rstrip(".,")

        # Try two-word combinations first (e.g. "po box", "north east")
        two_word = None
        if i + 1 < len(words):
            two_word_key = (wl + " " + words[i + 1].lower().rstrip(".,"))
            if two_word_key in UNIT_TYPES:
                result.append(UNIT_TYPES[two_word_key])
                i += 2
                continue
            if two_word_key in DIRECTIONALS:
                result.append(DIRECTIONALS[two_word_key])
                i += 2
                continue

        # Single word lookups
        if wl in STREET_TYPES:
            result.append(STREET_TYPES[wl])
        elif wl in UNIT_TYPES:
            result.append(UNIT_TYPES[wl])
        elif wl in DIRECTIONALS:
            result.append(DIRECTIONALS[wl])
        elif re.match(r"^\d+$", w):           # pure number
            result.append(w)
        elif re.match(r"^\d+(st|nd|rd|th)$", w, re.IGNORECASE):   # ordinal
            result.append(w.lower())
        elif re.match(r"^[a-zA-Z]\d+[a-zA-Z]?$", w):              # unit codes e.g. A2, B12C
            result.append(w.upper())
        elif re.match(r"^\d+[-–]\d+$", w):    # range e.g. 12-14
            result.append(w)
        else:
            result.append(w.title())
        i += 1

    return " ".join(result).upper()


def correct_city(val):
    val = _clean(val)
    if not val:
        return ""

    lower = val.lower()

    # 1. Exact match → canonical spelling
    if lower in _ALL_CITIES:
        return _ALL_CITIES[lower]

    if len(val) >= 3:
        from difflib import SequenceMatcher
        import math

        def _score(c):
            sim = SequenceMatcher(None, lower, c).ratio()
            pop_bonus = math.log10(_CITY_POPULATION.get(c, 0) + 1) * 0.045
            return sim + pop_bonus

        # 2a. Try large-city pool first (fast, high confidence)
        candidates = get_close_matches(lower, _LARGE_CITY_NAMES, n=5, cutoff=0.82)
        if candidates:
            best = max(candidates, key=_score)
            if SequenceMatcher(None, lower, best).ratio() >= 0.82:
                return _ALL_CITIES[best]

        # 2b. Try full pool (all 5k+ cities) — catches small towns
        candidates = get_close_matches(lower, _ALL_CITY_NAMES_FULL, n=5, cutoff=0.85)
        if candidates:
            best = max(candidates, key=_score)
            if SequenceMatcher(None, lower, best).ratio() >= 0.85:
                return _ALL_CITIES[best]

    # 3. Fallback — Title Case
    return val.title()


def correct_state(val, country_hint=""):
    val = _clean(val)
    if not val:
        return ""
    lower = val.lower()

    # 1. Exact match → abbreviation (already uppercase by definition)
    if lower in US_STATES:
        return US_STATES[lower]
    if lower in CA_PROVINCES:
        return CA_PROVINCES[lower]
    if lower in AU_STATES:
        return AU_STATES[lower]

    # 2. Short code (2–4 alpha chars) → uppercase
    if len(val) <= 4 and val.replace(" ", "").isalpha():
        return val.upper()

    # 3. Fuzzy match against all known state / province full names
    if len(val) > 4:
        hits = get_close_matches(lower, _ALL_STATE_NAMES, n=1, cutoff=0.78)
        if hits:
            return _STATE_FUZZY_INDEX[hits[0]]

    # 4. Fallback — all caps
    return val.upper()


def _pycountry_alpha2(c):
    """Return the ISO 3166-1 alpha-2 code for a pycountry country object."""
    return c.alpha_2


# Build a flat lookup: lowercase name variant → pycountry object
# Used for difflib fuzzy matching (covers misspellings)
_COUNTRY_NAME_INDEX: dict = {}
for _c in pycountry.countries:
    for _attr in ("name", "common_name", "official_name", "alpha_2", "alpha_3"):
        _v = getattr(_c, _attr, None)
        if _v:
            _COUNTRY_NAME_INDEX[_v.lower()] = _c
_ALL_COUNTRY_NAMES = list(_COUNTRY_NAME_INDEX.keys())

# Pre-built sorted list of country names (longest first) for greedy matching in company names
# Only names ≥ 4 chars to avoid false positives (e.g. "in", "us", "de")
_COUNTRY_WORDS_SORTED = sorted(
    [n for n in _ALL_COUNTRY_NAMES if len(n) >= 4],
    key=len, reverse=True
)

def _infer_country_from_company_name(name: str) -> str:
    """
    Scan a company / vendor name for an embedded country word.
    E.g. "RYDER TRUCK RENTAL CANADA LTD" → "CA"
         "DEUTSCHE BANK AG"              → "DE"
    Returns ISO alpha-2 code, or "" if nothing found.
    Only matches whole words (word-boundary check) to avoid false positives.
    """
    if not name:
        return ""
    name_lower = name.lower()
    for country_word in _COUNTRY_WORDS_SORTED:
        # Whole-word match only
        pat = r"\b" + re.escape(country_word) + r"\b"
        if re.search(pat, name_lower):
            obj = _COUNTRY_NAME_INDEX.get(country_word)
            if obj:
                return _pycountry_alpha2(obj).upper()
    return ""

# ── Global ISO 3166-2 subdivision index ──────────────────────────────────────
# Maps lowercase full-name → ISO subdivision code suffix (e.g. "california" → "CA")
# Built from pycountry for all countries + manual entries for common aliases.
_STATE_FUZZY_INDEX: dict[str, str] = {}
for _sub in pycountry.subdivisions:
    _code_suffix = _sub.code.split("-", 1)[-1]   # "US-CA" → "CA"
    _STATE_FUZZY_INDEX[_sub.name.lower()] = _code_suffix

# Manual aliases not in pycountry names
_STATE_FUZZY_INDEX.update({k: v for k, v in US_STATES.items()})
_STATE_FUZZY_INDEX.update({k: v for k, v in CA_PROVINCES.items()})
_STATE_FUZZY_INDEX.update({k: v for k, v in AU_STATES.items()})
_ALL_STATE_NAMES = list(_STATE_FUZZY_INDEX.keys())

# ── City fuzzy index ──────────────────────────────────────────────────────────
# Include cities with population > 5 000 to catch small towns.
# Two tiers keep the fuzzy match fast:
#   _ALL_CITIES / _CITY_POPULATION / _CITY_LOCATION — all cities > 5k
#   _LARGE_CITY_NAMES — only cities > 50k, used as the preferred fuzzy pool
#   (small-city fuzzy matching is only attempted when the large-city pool fails)
_gc = _gnc.GeonamesCache()
_ALL_CITIES: dict[str, str] = {}         # lowercase → canonical Title Case name
_CITY_POPULATION: dict[str, int] = {}    # lowercase → population (for tiebreaking)
_CITY_LOCATION: dict[str, tuple] = {}    # lowercase → (state_code, country_code)

for _city in _gc.get_cities().values():
    _pop = _city.get("population") or 0
    if _pop < 5_000:
        continue
    _name    = _city["name"]
    _key     = _name.lower()
    _state_c = (_city.get("subcountrycode") or "").strip().upper()
    _ctry_c  = (_city.get("countrycode")    or "").strip().upper()

    if _pop > _CITY_POPULATION.get(_key, 0):
        _ALL_CITIES[_key]      = _name
        _CITY_POPULATION[_key] = _pop
        _CITY_LOCATION[_key]   = (_state_c, _ctry_c)

    _alt = _name.replace(" City", "").replace(" city", "").strip()
    if _alt and _alt.lower() != _key:
        if _pop > _CITY_POPULATION.get(_alt.lower(), 0):
            _ALL_CITIES[_alt.lower()]      = _name
            _CITY_POPULATION[_alt.lower()] = _pop
            _CITY_LOCATION[_alt.lower()]   = (_state_c, _ctry_c)

_LARGE_CITY_NAMES = [k for k, v in _CITY_POPULATION.items() if v >= 50_000]
_ALL_CITY_NAMES_FULL = list(_ALL_CITIES.keys())   # all 5k+ cities

# Explicit common aliases not in geonamescache
_CITY_ALIASES = {
    "new york":    "New York City",
    "nyc":         "New York City",
    "la":          "Los Angeles",
    "sf":          "San Francisco",
    "dc":          "Washington",
    "philly":      "Philadelphia",
    "vegas":       "Las Vegas",
    "nola":        "New Orleans",
    "chi":         "Chicago",
}
_ALL_CITIES.update(_CITY_ALIASES)
_ALL_CITY_NAMES = _ALL_CITY_NAMES_FULL   # alias used by existing code


def correct_country(val):
    val = _clean(val)
    if not val:
        return ""

    lookup  = val.lower().strip(" .")
    no_dots = lookup.replace(".", "")
    stripped = no_dots.replace(" ", "")

    # 1. Manual alias map — fastest; handles abbreviations & native names
    if lookup in COUNTRY_MAP:
        return COUNTRY_MAP[lookup].upper()
    if no_dots in COUNTRY_MAP:
        return COUNTRY_MAP[no_dots].upper()

    # 2. ISO alpha-2 code  (e.g. "DE", "JP", "au")
    if len(stripped) == 2 and stripped.isalpha():
        c = pycountry.countries.get(alpha_2=stripped.upper())
        if c:
            return _pycountry_alpha2(c).upper()

    # 3. ISO alpha-3 code  (e.g. "DEU", "aus", "GBR")
    if len(stripped) == 3 and stripped.isalpha():
        c = pycountry.countries.get(alpha_3=stripped.upper())
        if c:
            return _pycountry_alpha2(c).upper()

    # 4. Exact ISO name / common_name / official_name match
    for attr, query in [
        ("name",          val.title()),
        ("common_name",   val.title()),
        ("official_name", val.title()),
        ("name",          val.upper()),
    ]:
        c = pycountry.countries.get(**{attr: query})
        if c:
            return _pycountry_alpha2(c).upper()

    # 5. pycountry token search
    try:
        results = pycountry.countries.search_fuzzy(val)
        if results:
            return _pycountry_alpha2(results[0]).upper()
    except LookupError:
        pass

    # 6. difflib edit-distance fuzzy match over all 249 country names
    hits = get_close_matches(lookup, _ALL_COUNTRY_NAMES, n=1, cutoff=0.72)
    if hits:
        return _pycountry_alpha2(_COUNTRY_NAME_INDEX[hits[0]]).upper()

    # 7. Fallback — all caps
    return val.upper()


def correct_postal_code(val, country_hint=""):
    val = _clean(val)
    if not val:
        return ""

    val = val.strip("'\"")
    stripped = val.replace(" ", "").upper()

    # ── Canadian postal code ────────────────────────────────────────────────
    # Format: A1A 1A1  (positions 0,2,3 are letters; 1,4,5 are digits)
    # Canadian codes never use D, F, I, O, Q, U — fix common I→1, O→0 OCR errors
    ca_raw = re.match(r"^([A-Za-z][0-9OoIi][A-Za-z])\s*([0-9OoIi][A-Za-z][0-9OoIi])$",
                      stripped)
    if ca_raw:
        def _fix_ca(s):
            # Positions that must be digits: index 1, 3, 5 (in compact form)
            s = list(s.replace(" ", "").upper())
            digit_pos = {1, 3, 5}
            for i in digit_pos:
                if i < len(s):
                    s[i] = s[i].replace("I", "1").replace("O", "0")
            return "".join(s)
        fixed = _fix_ca(stripped)
        return f"{fixed[:3]} {fixed[3:]}"

    # ── US ZIP+4 — valid 5+4 format ────────────────────────────────────────
    zip4 = re.match(r"^(\d{5})\s*[-–]\s*(\d{4})$", val)
    if zip4:
        return f"{zip4.group(1)}-{zip4.group(2)}"

    # ── Malformed US ZIP — 5+5 digits (PO Box merged with ZIP) ─────────────
    # e.g. "91117-26001" or "48278-78001" → keep only the 5-digit ZIP
    zip5plus = re.match(r"^(\d{5})[-–]\d{5,}$", val)
    if zip5plus:
        return zip5plus.group(1)

    # ── Pure numeric — strip spaces, preserve leading zeros ────────────────
    if stripped.isdigit():
        return stripped

    # ── UK postcode — ensure single space before inward code ───────────────
    uk_match = re.match(r"^([A-Z]{1,2}\d[A-Z\d]?)\s*(\d[A-Z]{2})$", stripped)
    if uk_match:
        return f"{uk_match.group(1)} {uk_match.group(2)}"

    # ── Netherlands 1234AB → 1234 AB ───────────────────────────────────────
    nl_match = re.match(r"^(\d{4})([A-Z]{2})$", stripped)
    if nl_match:
        return f"{nl_match.group(1)} {nl_match.group(2)}"

    return val.upper()


# ── Postal-code country detector ───────────────────────────────────────────────
# Only uses UNAMBIGUOUS formats — patterns that belong to exactly one country.
# Pure numeric codes (5-digit, 6-digit, 4-digit) are intentionally excluded
# because they are shared across many countries (US/DE/FR share 5-digit, etc.)
_POSTAL_PATTERNS: list[tuple] = [
    # Canadian  A1A 1A1 or A1A1A1  — alpha-numeric alternating, unique to Canada
    (re.compile(r"^[A-Z]\d[A-Z]\s?\d[A-Z]\d$"),               "CA"),
    # UK postcode  SW1A 2AA / EC1A 1BB  — letter(s)+digit(s)+optional-letter, digit+2-letters
    (re.compile(r"^[A-Z]{1,2}\d{1,2}[A-Z]?\s?\d[A-Z]{2}$"),  "GB"),
    # Netherlands  1234 AB  — 4 digits + exactly 2 uppercase letters
    (re.compile(r"^\d{4}\s?[A-Z]{2}$"),                        "NL"),
    # Irish Eircode  A65 F4E2  — letter+2digits, space, 4 alphanum
    (re.compile(r"^[A-Z]\d{2}\s?[A-Z0-9]{4}$"),               "IE"),
    # US ZIP+4  12345-6789  — only format using a hyphen between 5 and 4 digits
    (re.compile(r"^\d{5}-\d{4}$"),                             "US"),
    # Swiss  CH-NNNN  (sometimes written with prefix)
    (re.compile(r"^CH-\d{4}$"),                                "CH"),
    # German  DE-NNNNN (sometimes written with prefix)
    (re.compile(r"^DE-\d{5}$"),                                "DE"),
]

def detect_country_from_postal(postal: str):
    """Return ISO alpha-2 country code inferred from postal code format, or None."""
    if not postal:
        return None
    p = postal.strip().upper().replace(" ", "")
    p_spaced = postal.strip().upper()
    for pattern, country in _POSTAL_PATTERNS:
        if pattern.match(p) or pattern.match(p_spaced):
            return country
    return None


# ── Canadian FSA (Forward Sortation Area) → Province ─────────────────────────
# The first letter of a Canadian postal code uniquely identifies the province.
_CA_FSA_PROVINCE: dict[str, str] = {
    "A": "NL",   # Newfoundland and Labrador
    "B": "NS",   # Nova Scotia
    "C": "PE",   # Prince Edward Island
    "E": "NB",   # New Brunswick
    "G": "QC",   # Quebec (east)
    "H": "QC",   # Quebec (Montreal)
    "J": "QC",   # Quebec (west)
    "K": "ON",   # Ontario (east)
    "L": "ON",   # Ontario (central)
    "M": "ON",   # Ontario (Toronto)
    "N": "ON",   # Ontario (southwest)
    "P": "ON",   # Ontario (north)
    "R": "MB",   # Manitoba
    "S": "SK",   # Saskatchewan
    "T": "AB",   # Alberta
    "V": "BC",   # British Columbia
    "X": "NT",   # Northwest Territories / Nunavut
    "Y": "YT",   # Yukon
}

# US state codes — used to detect when a US state has been listed for a Canadian address
_US_STATE_CODES = set(US_STATES.values())


def infer_province_from_canadian_postal(postal: str):
    """Return the 2-letter Canadian province code from a postal code's FSA letter."""
    if not postal:
        return None
    first = postal.strip().upper()[0]
    return _CA_FSA_PROVINCE.get(first)


# ── US ZIP prefix (first 3 digits) → state code ───────────────────────────────
# Allows state validation/correction when a US ZIP code is present.
_ZIP_PREFIX_TO_STATE: dict[str, str] = {
    # Connecticut
    **{p: "CT" for p in ["060","061","062","063","064","065","066","067","068","069"]},
    # Maine
    **{p: "ME" for p in ["039","040","041","042","043","044","045","046","047","048","049"]},
    # Massachusetts
    **{p: "MA" for p in ["010","011","012","013","014","015","016","017","018","019",
                          "020","021","022","023","024","025","026","027"]},
    # New Hampshire
    **{p: "NH" for p in ["030","031","032","033","034","035","036","037","038"]},
    # Rhode Island
    **{p: "RI" for p in ["028","029"]},
    # Vermont
    **{p: "VT" for p in ["050","051","052","053","054","055","056","057","058","059"]},
    # New Jersey
    **{p: "NJ" for p in ["070","071","072","073","074","075","076","077","078","079",
                          "080","081","082","083","084","085","086","087","088","089"]},
    # New York
    **{p: "NY" for p in [f"{n:03d}" for n in range(100, 150)]},
    # Pennsylvania
    **{p: "PA" for p in [f"{n:03d}" for n in range(150, 197)]},
    # Delaware
    **{p: "DE" for p in ["197","198","199"]},
    # Washington DC
    **{p: "DC" for p in ["200","201","202","203","204","205"]},
    # Maryland
    **{p: "MD" for p in ["206","207","208","209","210","211","212","214","215","216","217","218","219"]},
    # Virginia
    **{p: "VA" for p in [f"{n:03d}" for n in range(220, 247)]},
    # West Virginia
    **{p: "WV" for p in [f"{n:03d}" for n in range(247, 269)]},
    # North Carolina
    **{p: "NC" for p in [f"{n:03d}" for n in range(270, 290)]},
    # South Carolina
    **{p: "SC" for p in [f"{n:03d}" for n in range(290, 300)]},
    # Georgia
    **{p: "GA" for p in [f"{n:03d}" for n in range(300, 320)] + ["398","399"]},
    # Florida
    **{p: "FL" for p in ["320","321","322","323","324","325","326","327","328","329",
                          "330","331","332","333","334","335","336","337","338","339",
                          "341","342","344","346","347","349"]},
    # Alabama
    **{p: "AL" for p in ["350","351","352","354","355","356","357","358","359",
                          "360","361","362","363","364","365","366","367","368","369"]},
    # Tennessee
    **{p: "TN" for p in ["370","371","372","373","374","375","376","377","378","379",
                          "380","381","382","383","384","385"]},
    # Mississippi
    **{p: "MS" for p in [f"{n:03d}" for n in range(386, 398)]},
    # Kentucky
    **{p: "KY" for p in ["400","401","402","403","404","405","406","407","408","409",
                          "410","411","412","413","414","415","416","417","418",
                          "420","421","422","423","424","425","426","427"]},
    # Ohio
    **{p: "OH" for p in [f"{n:03d}" for n in range(430, 459)]},
    # Indiana
    **{p: "IN" for p in [f"{n:03d}" for n in range(460, 480)]},
    # Michigan
    **{p: "MI" for p in [f"{n:03d}" for n in range(480, 500)]},
    # Iowa
    **{p: "IA" for p in ["500","501","502","503","504","505","506","507","508",
                          "510","511","512","513","514","515","516",
                          "520","521","522","523","524","525","526","527","528"]},
    # Wisconsin
    **{p: "WI" for p in ["530","531","532","534","535","537","538","539",
                          "540","541","542","543","544","545","546","547","548","549"]},
    # Minnesota
    **{p: "MN" for p in ["550","551","553","554","555","556","557","558","559",
                          "560","561","562","563","564","565","566","567"]},
    # South Dakota
    **{p: "SD" for p in ["570","571","572","573","574","575","576","577"]},
    # North Dakota
    **{p: "ND" for p in ["580","581","582","583","584","585","586","587","588"]},
    # Montana
    **{p: "MT" for p in [f"{n:03d}" for n in range(590, 600)]},
    # Illinois
    **{p: "IL" for p in [f"{n:03d}" for n in range(600, 630)]},
    # Missouri
    **{p: "MO" for p in ["630","631","633","634","635","636","637","638","639",
                          "640","641","644","645","646","647","648",
                          "650","651","652","653","654","655","656","657","658"]},
    # Kansas
    **{p: "KS" for p in ["660","661","662","664","665","666","667","668","669",
                          "670","671","672","673","674","675","676","677","678","679"]},
    # Nebraska
    **{p: "NE" for p in ["680","681","683","684","685","686","687","688","689",
                          "690","691","692","693"]},
    # Louisiana
    **{p: "LA" for p in ["700","701","703","704","705","706","707","708",
                          "710","711","712","713","714"]},
    # Arkansas
    **{p: "AR" for p in ["716","717","718","719","720","721","722","723","724","725","726","727","728","729"]},
    # Oklahoma
    **{p: "OK" for p in ["730","731","734","735","736","737","738","739",
                          "740","741","743","744","745","746","747","748","749"]},
    # Texas
    **{p: "TX" for p in [f"{n:03d}" for n in range(750, 800)] + ["885"]},
    # Colorado
    **{p: "CO" for p in [f"{n:03d}" for n in range(800, 817)]},
    # Wyoming
    **{p: "WY" for p in ["820","821","822","823","824","825","826","827","828","829","830","831"]},
    # Idaho
    **{p: "ID" for p in ["832","833","834","835","836","837","838"]},
    # Utah
    **{p: "UT" for p in ["840","841","842","843","844","845","846","847"]},
    # Arizona
    **{p: "AZ" for p in ["850","851","852","853","855","856","857","859",
                          "860","863","864","865"]},
    # New Mexico
    **{p: "NM" for p in ["870","871","872","873","874","875","877","878","879",
                          "880","881","882","883","884"]},
    # Nevada
    **{p: "NV" for p in ["889","890","891","893","894","895","897","898"]},
    # Hawaii
    **{p: "HI" for p in ["967","968"]},
    # California
    **{p: "CA" for p in [f"{n:03d}" for n in
                          list(range(900, 909)) + list(range(910, 962)) if n not in [909,929]]},
    # Oregon
    **{p: "OR" for p in [f"{n:03d}" for n in range(970, 980)]},
    # Washington
    **{p: "WA" for p in ["980","981","982","983","984","985","986","988","989",
                          "990","991","992","993","994"]},
    # Alaska
    **{p: "AK" for p in [f"{n:03d}" for n in range(995, 1000)]},
    # Puerto Rico
    **{p: "PR" for p in ["006","007","008","009"]},
    # Guam
    "969": "GU",
    # APO/FPO
    **{p: "AE" for p in ["090","091","092","093","094","095","096","097","098"]},
}


def infer_us_state_from_zip(postal: str) -> str:
    """
    Return the expected US state code from a 5-digit ZIP code's first 3 digits.
    Returns "" if the ZIP is not a recognised US format.
    """
    if not postal:
        return ""
    digits = postal.strip().replace("-", "")[:5]
    if len(digits) < 3 or not digits.isdigit():
        return ""
    return _ZIP_PREFIX_TO_STATE.get(digits[:3], "")



CORRECTORS = {
    "address_line_1": correct_address_line,
    "address_line_2": correct_address_line,
    "address_line_3": correct_address_line,
    "city":           correct_city,
    "state":          correct_state,
    "country":        correct_country,
    "postal_code":    correct_postal_code,
}

DISPLAY_LABELS = {
    "address_line_1": "Address Line 1",
    "address_line_2": "Address Line 2",
    "address_line_3": "Address Line 3",
    "city":           "City",
    "state":          "State",
    "country":        "Country",
    "postal_code":    "Postal Code",
}

# ──────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT FORMATTING
# ──────────────────────────────────────────────────────────────────────────────

_HDR_FILL_ORIG = PatternFill("solid", fgColor="1F4E79")   # dark navy – original
_HDR_FILL_CORR = PatternFill("solid", fgColor="375623")   # dark green – corrected
_HDR_FONT      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_DATA_FONT     = Font(name="Arial", size=10)
_CHANGED_FILL  = PatternFill("solid", fgColor="E2EFDA")   # light green – cell changed
_ALT_FILL      = PatternFill("solid", fgColor="F2F2F2")   # light grey – alt rows
_BORDER_SIDE   = Side(style="thin", color="BFBFBF")
_CELL_BORDER   = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE,  bottom=_BORDER_SIDE,
)


def _write_excel(result_df, orig_cols, corrected_col_map, output_path, col_map):
    wb = Workbook()
    ws = wb.active
    ws.title = "Corrected Addresses"

    all_cols = list(result_df.columns)
    orig_set  = set(orig_cols)
    corr_set  = set(corrected_col_map.values())

    # ── Header row ──────────────────────────────────────────────────────────
    for c_idx, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font      = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _CELL_BORDER
        if col_name in corr_set:
            cell.fill = _HDR_FILL_CORR
        else:
            cell.fill = _HDR_FILL_ORIG

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Data rows ───────────────────────────────────────────────────────────
    # Build reverse map: corrected_col → original_col for change detection
    rev_corr = {v: k for k, v in corrected_col_map.items()}

    for r_idx, row in enumerate(result_df.itertuples(index=False), start=2):
        alt = (r_idx % 2 == 0)
        for c_idx, col_name in enumerate(all_cols, start=1):
            val = getattr(row, col_name.replace(" ", "_").replace("/", "_")
                          .replace("(", "").replace(")", ""), None)
            # Fallback using positional index
            val = result_df.iloc[r_idx - 2, c_idx - 1]

            cell = ws.cell(row=r_idx, column=c_idx, value=val if val != "" else None)
            cell.font      = _DATA_FONT
            cell.alignment = Alignment(vertical="center")
            cell.border    = _CELL_BORDER

            # Highlight changed corrected cells
            if col_name in corr_set:
                orig_col = rev_corr[col_name]
                orig_val = result_df.iloc[r_idx - 2][orig_col] if orig_col else ""
                if str(val or "").strip() != str(orig_val or "").strip():
                    cell.fill = _CHANGED_FILL
                elif alt:
                    cell.fill = _ALT_FILL
            elif alt:
                cell.fill = _ALT_FILL

    # ── Column widths ────────────────────────────────────────────────────────
    for c_idx, col_name in enumerate(all_cols, start=1):
        col_vals = result_df.iloc[:, c_idx - 1].astype(str)
        max_len  = max(col_vals.str.len().max(), len(col_name))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 4, 40)

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Address Correction Summary"
    ws2["A1"].font = Font(bold=True, name="Arial", size=13)
    ws2.merge_cells("A1:C1")

    ws2["A3"] = "Total Rows Processed"
    ws2["B3"] = len(result_df)
    ws2["A4"] = "Columns Detected & Corrected"
    ws2["B4"] = len(corrected_col_map)

    ws2["A6"] = "Field"
    ws2["B6"] = "Original Column"
    ws2["C6"] = "Corrections Made"
    for cell in (ws2["A6"], ws2["B6"], ws2["C6"]):
        cell.font = Font(bold=True, name="Arial", size=10)
        cell.fill = _HDR_FILL_ORIG
        cell.font = _HDR_FONT

    row_num = 7
    for field, orig_col in col_map.items():
        if orig_col is None:
            continue
        if field not in DISPLAY_LABELS:
            continue  # hint-only fields (e.g. company_name) have no corrected column
        corr_col = corrected_col_map.get(f"Corrected {DISPLAY_LABELS[field]}")
        if corr_col is None:
            continue
        changed = (
            result_df[orig_col].astype(str).str.strip()
            != result_df[corr_col].astype(str).str.strip()
        ).sum()
        ws2.cell(row=row_num, column=1, value=DISPLAY_LABELS[field]).font = _DATA_FONT
        ws2.cell(row=row_num, column=2, value=orig_col).font              = _DATA_FONT
        ws2.cell(row=row_num, column=3, value=int(changed)).font          = _DATA_FONT
        row_num += 1

    for col in ("A", "B", "C"):
        ws2.column_dimensions[col].width = 28

    wb.save(output_path)


# ──────────────────────────────────────────────────────────────────────────────
# AUTO-FIX  (reusable by both process_file and the Streamlit app)
# ──────────────────────────────────────────────────────────────────────────────

# State code → country mapping (codes that are UNAMBIGUOUS to one country)
# Only include codes that cannot belong to multiple countries
_STATE_CODE_TO_COUNTRY: dict[str, str] = {}

# Canadian province codes → CA  (none overlap with US states)
for _code in CA_PROVINCE_CODES:
    _STATE_CODE_TO_COUNTRY[_code] = "CA"

# Australian state codes → AU
# NT overlaps with CA's Northwest Territories — skip it to avoid false positives
for _code in AU_STATE_CODES - {"NT", "SA"}:
    _STATE_CODE_TO_COUNTRY[_code] = "AU"

# US state codes only added where they don't clash with AU/CA
# (US has no overlap with CA province codes; AU "SA" clashes with nothing but keep cautious)
# We do NOT add US state codes here — US is the default assumption already;
# we only want to OVERRIDE when the state proves a *different* country.


def apply_autofix(result: "pd.DataFrame", col_map: dict) -> None:
    """
    In-place: fix Corrected Country and Corrected State for every row using
    multiple evidence signals, strongest first:

      1. Postal code format  — unambiguous country patterns (CA A1A1A1, GB SW1A2AA, etc.)
      2. State/region code   — if corrected state is an unambiguous province/territory code
                               (e.g. ON/AB/QC → CA;  NSW/VIC/QLD → AU)
      3. Company/vendor name — country word embedded in name (e.g. "CANADA LTD" → CA)
      4. Province from FSA   — infer Canadian province from postal FSA first letter
    """
    postal_col   = col_map.get("postal_code")
    name_col     = col_map.get("company_name")
    city_col_orig = col_map.get("city")          # original city column (for future use)
    corr_country  = "Corrected Country"
    corr_state    = "Corrected State"
    corr_postal   = "Corrected Postal Code"

    if corr_country not in result.columns:
        return

    _null_upper = {s.upper() for s in _NULL_PLACEHOLDERS}

    def _fix_row(row):
        postal    = str(row.get(corr_postal) or (row.get(postal_col) if postal_col else "") or "").strip()
        country   = str(row.get(corr_country) or "").strip()
        raw_state = str(row.get(corr_state)   or "").strip() if corr_state in row.index else ""
        state     = "" if raw_state.upper() in _null_upper else raw_state

        postal_inferred = detect_country_from_postal(postal) if postal_col else ""

        # ── Signal 1: postal code (strongest) ────────────────────────────────
        if postal_inferred and postal_inferred != country:
            country = postal_inferred

        # ── Signal 2: state/region code unambiguously identifies a country ───
        # Only fires when postal didn't already give us a confident answer,
        # OR when postal confirmed the same country (so state just fills in the gap).
        if state:
            state_country = _STATE_CODE_TO_COUNTRY.get(state.upper(), "")
            if state_country and state_country != country:
                # Don't override a confident postal-code answer with a weaker state hint
                if not postal_inferred:
                    country = state_country

        # ── Signal 2b: US state code → assume US when country still unknown ──
        # US codes are intentionally absent from _STATE_CODE_TO_COUNTRY (to avoid
        # false overrides), so we check _US_STATE_CODES separately here.
        if not country and state and state.upper() in _US_STATE_CODES:
            country = "US"

        # ── Signal 3: company/vendor name ────────────────────────────────────
        if name_col:
            company = str(row.get(name_col) or "").strip()
            name_country = _infer_country_from_company_name(company)
            if name_country and name_country != country and not postal_inferred:
                # also don't override if state already gave us an answer
                if not _STATE_CODE_TO_COUNTRY.get(state.upper(), ""):
                    country = name_country

        # ── Signal 4: Canadian province from FSA first letter ─────────────────
        if country == "CA":
            prov = infer_province_from_canadian_postal(postal)
            if prov and (not state or state in _US_STATE_CODES or state.upper() in _null_upper):
                state = prov

        # ── Signal 5: US state from ZIP prefix — authoritative override ─────
        # Fires when: country is US, blank, OR state is a known US code.
        zip_state = infer_us_state_from_zip(postal)
        _is_us    = country == "US" or country == "" or state.upper() in _US_STATE_CODES
        if zip_state and _is_us:
            state   = zip_state
            country = "US"

        # ── Signal 6: Australian state — infer country if not already CA/GB ──
        if state.upper() in AU_STATE_CODES and country not in ("CA", "GB", "US"):
            if not postal_inferred:
                country = "AU"

        return pd.Series({corr_country: country, corr_state: state})

    fixed = result.apply(_fix_row, axis=1)
    result[corr_country] = fixed[corr_country]
    if corr_state in result.columns:
        mask = fixed[corr_state] != ""
        result.loc[mask, corr_state] = fixed.loc[mask, corr_state]


# ──────────────────────────────────────────────────────────────────────────────
# MAIN PROCESS
# ──────────────────────────────────────────────────────────────────────────────

def process_file(input_path, output_path):
    ext = os.path.splitext(input_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(input_path, dtype=str)
    elif ext == ".csv":
        df = pd.read_csv(input_path, dtype=str)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Use .csv, .xls, or .xlsx")

    df = df.fillna("")

    col_map = _detect_columns(df.columns)

    # Print detected columns
    print("\nDetected columns:")
    for field, col in col_map.items():
        status = f'"{col}"' if col else "NOT FOUND"
        label = DISPLAY_LABELS.get(field, field.replace("_", " ").title())
        print(f"  {label:20s} -> {status}")

    # Apply corrections
    result = df.copy()
    corrected_col_map = {}  # "Corrected X" → series

    for field, orig_col in col_map.items():
        if orig_col is None:
            continue
        # company_name is a hint-only field — we don't produce a corrected column for it
        if field not in CORRECTORS:
            continue
        label = f"Corrected {DISPLAY_LABELS[field]}"
        result[label] = df[orig_col].apply(CORRECTORS[field])
        corrected_col_map[label] = label

    # ── Auto-fix country & state using postal code + company name evidence ──
    apply_autofix(result, col_map)

    # Write output
    out_ext = os.path.splitext(output_path)[1].lower()
    if out_ext == ".csv":
        result.to_csv(output_path, index=False)
        print(f"\nDone. Output written to: {output_path}")
    else:
        _write_excel(result, list(df.columns), corrected_col_map, output_path, col_map)
        print(f"\nDone. Output written to: {output_path}")

    # Print summary
    print(f"Rows processed     : {len(df)}")
    print(f"Corrected columns  : {len(corrected_col_map)}")
    total_changes = 0
    for field, orig_col in col_map.items():
        if orig_col is None:
            continue
        if field not in DISPLAY_LABELS:
            continue  # hint-only fields have no corrected column
        label = f"Corrected {DISPLAY_LABELS[field]}"
        changed = (
            result[orig_col].astype(str).str.strip()
            != result[label].astype(str).str.strip()
        ).sum()
        total_changes += changed
        if changed:
            print(f"  {DISPLAY_LABELS[field]:20s}: {changed} cell(s) corrected")
    print(f"Total corrections  : {total_changes}")

    return result


# ──────────────────────────────────────────────────────────────────────────────
# CONFIDENCE SCORING
# ──────────────────────────────────────────────────────────────────────────────



if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        print("Example:")
        print("  python address_corrector.py addresses.csv   corrected.xlsx")
        print("  python address_corrector.py addresses.xlsx  corrected.xlsx")
        sys.exit(1)

    input_file  = sys.argv[1]
    output_file = sys.argv[2]

    if not os.path.exists(input_file):
        print(f"Error: Input file not found: {input_file}")
        sys.exit(1)

    process_file(input_file, output_file)
